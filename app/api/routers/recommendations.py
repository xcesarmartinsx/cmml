"""
app/api/routers/recommendations.py
===================================
Router FastAPI para o dashboard de ofertas.

Endpoints:
  GET  /api/recommendations/batches              — lista todos os batches com resumo
  GET  /api/recommendations/offers               — ofertas do batch mais recente (ou específico)
  GET  /api/recommendations/offers/export         — exporta ofertas em CSV (sem paginação, telefone completo)
  GET  /api/recommendations/offers/export-feedback — exporta Excel para feedback manual
  GET  /api/recommendations/summary              — KPIs + funil de scores do batch atual
  GET  /api/recommendations/product-lifecycle     — Top 50 produtos com ciclo de vida e estatísticas
  GET  /api/recommendations/lifecycle             — ciclo de vida de todos os produtos (com busca e filtro)
  GET  /api/recommendations/lifecycle/stats       — estatísticas agregadas por tier
  GET  /api/recommendations/feedback/summary      — métricas de conversão
  GET  /api/recommendations/feedback/runs         — histórico de runs de feedback
  POST /api/recommendations/feedback/run          — dispara cross-reference manual (admin)
  POST /api/recommendations/feedback/import       — importa feedback via Excel (admin)
"""

import csv
import io
import os
from datetime import date
from typing import Optional

import psycopg2
import psycopg2.extras
from fastapi import APIRouter, Query, Depends, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse

from deps import require_admin, get_current_user, get_current_user_info

router = APIRouter()


def _get_db():
    return psycopg2.connect(
        host=os.getenv("PG_HOST", "postgres"),
        port=int(os.getenv("PG_PORT", "5432")),
        dbname=os.getenv("PG_DB", "reco"),
        user=os.getenv("PG_USER", "reco"),
        password=os.environ["PG_PASSWORD"],
    )


@router.get("/api/recommendations/batches")
def get_batches():
    """
    Lista todos os batches de recomendações em ordem cronológica reversa.

    Retorna por batch: offer_batch_id, generated_at, expires_at,
    n_offers (total de linhas), n_customers (clientes únicos), n_sent (WhatsApp).
    """
    conn = _get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""
                SELECT
                    offer_batch_id::text,
                    MIN(generated_at)  AS generated_at,
                    MIN(expires_at)    AS expires_at,
                    COUNT(*)           AS n_offers,
                    COUNT(DISTINCT customer_id) AS n_customers,
                    COUNT(sent_via_whatsapp_at) AS n_sent
                FROM reco.offers
                GROUP BY offer_batch_id
                ORDER BY MIN(generated_at) DESC
            """)
            rows = cur.fetchall()
            result = []
            for r in rows:
                row = dict(r)
                if row.get("generated_at"):
                    row["generated_at"] = row["generated_at"].isoformat()
                if row.get("expires_at"):
                    row["expires_at"] = row["expires_at"].isoformat()
                row["n_offers"]    = int(row["n_offers"])
                row["n_customers"] = int(row["n_customers"])
                row["n_sent"]      = int(row["n_sent"])
                result.append(row)
            return result
    finally:
        conn.close()


@router.get("/api/recommendations/offers")
def get_offers(
    batch_id: Optional[str] = Query(None, description="UUID do batch; padrão: mais recente"),
    strategy: Optional[str] = Query(None, description="Filtrar por estratégia: 'modelo_a_ranker' ou 'modelo_b_colaborativo'"),
    sort:     str = Query("score_desc", description="Ordenação: score_desc, score_asc, price_desc, price_asc"),
    limit:    int = Query(100, ge=1, le=10000),
    offset:   int = Query(0, ge=0),
):
    """
    Retorna as ofertas do batch especificado (ou do mais recente).

    Para cada oferta inclui: nome do cliente, telefone, produto, score,
    rank e rastreabilidade de envio WhatsApp.

    O telefone exibido é o celular (mobile) se disponível; caso contrário,
    o telefone fixo (phone). A preferência é sempre pelo número de celular.
    """
    conn = _get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            # Determina o batch_id alvo
            if batch_id is None:
                cur.execute("""
                    SELECT offer_batch_id::text
                    FROM reco.offers
                    ORDER BY generated_at DESC
                    LIMIT 1
                """)
                row = cur.fetchone()
                if row is None:
                    return []
                batch_id = row["offer_batch_id"]

            params = [batch_id]
            strategy_filter = ""
            if strategy:
                strategy_filter = "AND o.strategy = %s"
                params.append(strategy)

            query = f"""
                SELECT
                    o.offer_id,
                    o.offer_batch_id::text,
                    o.customer_id,
                    sc.name                                   AS customer_name,
                    -- Valida telefone: deve ter ao menos 6 dígitos numéricos para ser útil.
                    -- Filtra valores inválidos do ERP como '(  )    -', '0', '0000000000000'.
                    NULLIF(
                        CASE
                            WHEN LENGTH(REGEXP_REPLACE(COALESCE(sc.mobile, sc.phone), '[^0-9]', '', 'g')) >= 6
                            THEN COALESCE(sc.mobile, sc.phone)
                            ELSE NULL
                        END,
                    '') AS contact,
                    CASE
                        WHEN wc.has_whatsapp = TRUE  THEN 'whatsapp'
                        WHEN wc.has_whatsapp = FALSE THEN 'mobile_no_wpp'
                        WHEN LENGTH(REGEXP_REPLACE(COALESCE(sc.mobile, sc.phone), '[^0-9]', '', 'g')) = 11
                         AND SUBSTRING(REGEXP_REPLACE(COALESCE(sc.mobile, sc.phone), '[^0-9]', '', 'g'), 3, 1) = '9'
                        THEN 'mobile'
                        WHEN LENGTH(REGEXP_REPLACE(COALESCE(sc.mobile, sc.phone), '[^0-9]', '', 'g')) >= 10
                        THEN 'landline'
                        ELSE 'unknown'
                    END AS phone_type,
                    sc.mobile,
                    sc.phone,
                    o.product_id,
                    cp.description                            AS product_name,
                    o.strategy,
                    o.score,
                    ROUND((o.score * 100)::numeric, 0)::int  AS score_pct,
                    ROUND((COALESCE(o.score_raw, o.score) * 100)::numeric, 0)::int AS score_raw_pct,
                    o.rank,
                    o.generated_at,
                    o.expires_at,
                    o.sent_via_whatsapp_at,
                    -- Última vez que este cliente comprou especificamente este produto.
                    -- NULL = nunca comprou (produto novo para o cliente).
                    lp.last_purchase_date,
                    COALESCE(pp.avg_unit_price, 0)::float AS avg_unit_price,
                    oo.converted,
                    oo.conversion_date,
                    oo.conversion_source
                FROM reco.offers o
                JOIN stg.customers sc
                    ON sc.customer_id_src = o.customer_id
                   AND sc.source_system   = 'sqlserver_gp'
                JOIN cur.products cp
                    ON cp.product_id    = o.product_id
                   AND cp.source_system = 'sqlserver_gp'
                LEFT JOIN reco.whatsapp_cache wc
                    ON wc.phone_number = '55' || REGEXP_REPLACE(COALESCE(sc.mobile, sc.phone), '[^0-9]', '', 'g')
                   AND wc.validated_at > now() - interval '30 days'
                LEFT JOIN cur.customer_product_last_purchase lp
                    ON lp.customer_id = o.customer_id
                   AND lp.product_id  = o.product_id
                LEFT JOIN cur.product_avg_price pp
                    ON pp.product_id = o.product_id
                LEFT JOIN (
                    SELECT DISTINCT ON (offer_id) offer_id, converted, conversion_date, conversion_source
                    FROM reco.offer_outcomes
                    ORDER BY offer_id, matched_at DESC
                ) oo ON oo.offer_id = o.offer_id
                WHERE o.offer_batch_id = %s::uuid
                  {strategy_filter}
                ORDER BY {
                    {"score_desc": "o.score DESC",
                     "score_asc":  "o.score ASC",
                     "price_desc": "COALESCE(pp.avg_unit_price, 0) DESC",
                     "price_asc":  "COALESCE(pp.avg_unit_price, 0) ASC",
                    }.get(sort, "o.score DESC")
                }, o.customer_id, o.rank
                LIMIT %s OFFSET %s
            """
            params.extend([limit, offset])
            cur.execute(query, params)
            rows = cur.fetchall()

            result = []
            for r in rows:
                row = dict(r)
                row["offer_id"]   = int(row["offer_id"])
                row["customer_id"] = int(row["customer_id"])
                row["product_id"]  = int(row["product_id"])
                row["score"]       = float(row["score"])
                row["score_pct"]   = int(row["score_pct"]) if row["score_pct"] is not None else 0
                row["rank"]        = int(row["rank"])
                if row.get("generated_at"):
                    row["generated_at"] = row["generated_at"].isoformat()
                if row.get("expires_at"):
                    row["expires_at"] = row["expires_at"].isoformat()
                if row.get("sent_via_whatsapp_at"):
                    row["sent_via_whatsapp_at"] = row["sent_via_whatsapp_at"].isoformat()
                if row.get("last_purchase_date"):
                    row["last_purchase_date"] = row["last_purchase_date"].isoformat()
                row["avg_unit_price"] = float(row["avg_unit_price"]) if row.get("avg_unit_price") else 0
                row["converted"] = row.get("converted")  # True/False/None
                if row.get("conversion_date"):
                    row["conversion_date"] = row["conversion_date"].isoformat()
                row["conversion_source"] = row.get("conversion_source")
                result.append(row)
            return result
    finally:
        conn.close()


def _format_phone(raw: str | None) -> str:
    """
    Formata o número de telefone para exibição legível na exportação.

    Exemplos:
      "11912345678"      ->  "(11) 91234-5678"
      "(11) 91234-5678"  ->  "(11) 91234-5678"
      None / ""          ->  ""
    """
    if not raw or not raw.strip():
        return ""
    digits = "".join(c for c in raw if c.isdigit())
    if len(digits) < 10:
        return raw.strip()
    if len(digits) == 11:
        return f"({digits[:2]}) {digits[2:7]}-{digits[7:]}"
    if len(digits) == 10:
        return f"({digits[:2]}) {digits[2:6]}-{digits[6:]}"
    return raw.strip()


@router.get("/api/recommendations/offers/export")
def export_offers(
    batch_id: Optional[str] = Query(None, description="UUID do batch; padrão: mais recente"),
    strategy: Optional[str] = Query(None, description="Filtrar por estratégia: 'modelo_a_ranker' ou 'modelo_b_colaborativo'"),
):
    """
    Exporta todas as ofertas do batch especificado (ou do mais recente) em CSV.

    Retorna StreamingResponse com media_type='text/csv'.
    O telefone é exibido completo para uso operacional das vendedoras.
    O endpoint herda a autenticação JWT exigida pelo router em main.py.
    """
    conn = _get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            # Determina o batch_id alvo
            if batch_id is None:
                cur.execute("""
                    SELECT offer_batch_id::text
                    FROM reco.offers
                    ORDER BY generated_at DESC
                    LIMIT 1
                """)
                row = cur.fetchone()
                if row is None:
                    # Retorna CSV vazio com cabeçalho
                    buf = io.StringIO()
                    writer = csv.writer(buf)
                    writer.writerow(["#", "cliente", "produto", "valor_unitario", "chance_pct", "modelo", "ultima_compra", "contato", "tipo_telefone", "gerado_em"])
                    buf.seek(0)
                    filename = f"ofertas_{date.today().isoformat()}.csv"
                    return StreamingResponse(
                        iter([buf.read()]),
                        media_type="text/csv",
                        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""},
                    )
                batch_id = row["offer_batch_id"]

            params = [batch_id]
            strategy_filter = ""
            if strategy:
                strategy_filter = "AND o.strategy = %s"
                params.append(strategy)

            query = f"""
                SELECT
                    o.offer_id,
                    o.rank,
                    sc.name                                   AS customer_name,
                    NULLIF(
                        CASE
                            WHEN LENGTH(REGEXP_REPLACE(COALESCE(sc.mobile, sc.phone), '[^0-9]', '', 'g')) >= 6
                            THEN COALESCE(sc.mobile, sc.phone)
                            ELSE NULL
                        END,
                    '') AS contact,
                    CASE
                        WHEN wc.has_whatsapp = TRUE  THEN 'whatsapp'
                        WHEN wc.has_whatsapp = FALSE THEN 'mobile_no_wpp'
                        WHEN LENGTH(REGEXP_REPLACE(COALESCE(sc.mobile, sc.phone), '[^0-9]', '', 'g')) = 11
                         AND SUBSTRING(REGEXP_REPLACE(COALESCE(sc.mobile, sc.phone), '[^0-9]', '', 'g'), 3, 1) = '9'
                        THEN 'mobile'
                        WHEN LENGTH(REGEXP_REPLACE(COALESCE(sc.mobile, sc.phone), '[^0-9]', '', 'g')) >= 10
                        THEN 'landline'
                        ELSE 'unknown'
                    END AS phone_type,
                    cp.description                            AS product_name,
                    o.strategy,
                    ROUND((o.score * 100)::numeric, 0)::int  AS score_pct,
                    COALESCE(pp.avg_unit_price, 0)::float     AS avg_unit_price,
                    o.generated_at,
                    lp.last_purchase_date
                FROM reco.offers o
                JOIN stg.customers sc
                    ON sc.customer_id_src = o.customer_id
                   AND sc.source_system   = 'sqlserver_gp'
                JOIN cur.products cp
                    ON cp.product_id    = o.product_id
                   AND cp.source_system = 'sqlserver_gp'
                LEFT JOIN reco.whatsapp_cache wc
                    ON wc.phone_number = '55' || REGEXP_REPLACE(COALESCE(sc.mobile, sc.phone), '[^0-9]', '', 'g')
                   AND wc.validated_at > now() - interval '30 days'
                LEFT JOIN cur.customer_product_last_purchase lp
                    ON lp.customer_id = o.customer_id
                   AND lp.product_id  = o.product_id
                LEFT JOIN cur.product_avg_price pp
                    ON pp.product_id = o.product_id
                WHERE o.offer_batch_id = %s::uuid
                  {strategy_filter}
                ORDER BY o.rank ASC
            """
            cur.execute(query, params)
            rows = cur.fetchall()

        # Monta o CSV em memória e faz streaming
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(["#", "cliente", "produto", "valor_unitario", "chance_pct", "modelo", "ultima_compra", "contato", "tipo_telefone", "gerado_em"])

        strategy_labels = {
            "modelo_a_ranker":       "Modelo A",
            "modelo_b_colaborativo": "Modelo B",
        }

        for rank_global, r in enumerate(rows, start=1):
            row = dict(r)

            last_purchase = ""
            if row.get("last_purchase_date"):
                d = row["last_purchase_date"]
                last_purchase = d.strftime("%d/%m/%Y") if hasattr(d, "strftime") else str(d)

            generated_at = ""
            if row.get("generated_at"):
                g = row["generated_at"]
                generated_at = g.strftime("%d/%m/%Y %H:%M") if hasattr(g, "strftime") else str(g)

            phone_type_label = {
                "whatsapp": "WPP \u2713",
                "mobile": "Celular",
                "mobile_no_wpp": "Sem WPP",
                "landline": "Fixo",
            }.get(row.get("phone_type", ""), "")

            writer.writerow([
                rank_global,
                row.get("customer_name") or "",
                row.get("product_name") or "",
                float(row["avg_unit_price"]) if row.get("avg_unit_price") else 0,
                int(row["score_pct"]) if row.get("score_pct") is not None else 0,
                strategy_labels.get(row.get("strategy", ""), row.get("strategy", "")),
                last_purchase,
                _format_phone(row.get("contact")),
                phone_type_label,
                generated_at,
            ])

        buf.seek(0)
        filename = f"ofertas_{date.today().isoformat()}.csv"
        return StreamingResponse(
            iter([buf.read()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=\"{filename}\""},
        )
    finally:
        conn.close()


@router.get("/api/recommendations/summary")
def get_summary():
    """
    KPIs consolidados + funil de scores do batch mais recente.

    Retorna: total_offers, n_customers, avg_score_pct, pct_bought_before,
    offer_batch_id, generated_at, n_modelo_a, n_modelo_b e lista funnel[]
    com distribuição de ofertas por faixa de score separada por modelo.
    """
    conn = _get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""
                WITH batch AS (
                    SELECT offer_batch_id
                    FROM   reco.offers
                    ORDER  BY generated_at DESC
                    LIMIT  1
                ),
                purchases AS (
                    SELECT DISTINCT customer_id, product_id
                    FROM   cur.order_items
                ),
                offers AS (
                    SELECT
                        o.*,
                        (p.customer_id IS NOT NULL) AS bought_before
                    FROM reco.offers o
                    LEFT JOIN purchases p
                           ON p.customer_id = o.customer_id
                          AND p.product_id  = o.product_id
                    WHERE o.offer_batch_id = (SELECT offer_batch_id FROM batch)
                )
                SELECT
                    offer_batch_id::text,
                    MIN(generated_at)                                                           AS generated_at,
                    COUNT(*)::int                                                               AS total_offers,
                    COUNT(DISTINCT customer_id)::int                                            AS n_customers,
                    ROUND(AVG(score) * 100)::int                                                AS avg_score_pct,
                    ROUND(AVG(CASE WHEN bought_before THEN 1.0 ELSE 0.0 END) * 100, 1)::float  AS pct_bought_before,
                    COUNT(*) FILTER (WHERE score >= 0.8)::int                                   AS score_80plus,
                    COUNT(*) FILTER (WHERE score >= 0.6 AND score < 0.8)::int                   AS score_60_79,
                    COUNT(*) FILTER (WHERE score >= 0.4 AND score < 0.6)::int                   AS score_40_59,
                    COUNT(*) FILTER (WHERE score  < 0.4)::int                                   AS score_below_40,
                    COUNT(*) FILTER (WHERE strategy = 'modelo_a_ranker')::int                   AS n_modelo_a,
                    COUNT(*) FILTER (WHERE strategy = 'modelo_b_colaborativo')::int             AS n_modelo_b,
                    COUNT(*) FILTER (WHERE strategy = 'modelo_a_ranker' AND score >= 0.8)::int              AS a_80plus,
                    COUNT(*) FILTER (WHERE strategy = 'modelo_a_ranker' AND score >= 0.6 AND score < 0.8)::int AS a_60_79,
                    COUNT(*) FILTER (WHERE strategy = 'modelo_a_ranker' AND score >= 0.4 AND score < 0.6)::int AS a_40_59,
                    COUNT(*) FILTER (WHERE strategy = 'modelo_a_ranker' AND score  < 0.4)::int              AS a_below_40,
                    COUNT(*) FILTER (WHERE strategy = 'modelo_b_colaborativo' AND score >= 0.8)::int              AS b_80plus,
                    COUNT(*) FILTER (WHERE strategy = 'modelo_b_colaborativo' AND score >= 0.6 AND score < 0.8)::int AS b_60_79,
                    COUNT(*) FILTER (WHERE strategy = 'modelo_b_colaborativo' AND score >= 0.4 AND score < 0.6)::int AS b_40_59,
                    COUNT(*) FILTER (WHERE strategy = 'modelo_b_colaborativo' AND score  < 0.4)::int              AS b_below_40
                FROM offers
                GROUP BY offer_batch_id
            """)
            row = cur.fetchone()
            if row is None:
                return {}

            r = dict(row)
            if r.get("generated_at"):
                r["generated_at"] = r["generated_at"].isoformat()

            total = r["total_offers"] or 1
            r["funnel"] = [
                {
                    "label": "≥ 80%",  "color": "#16a34a",
                    "total": r.pop("score_80plus"),
                    "modelo_a": r.pop("a_80plus"),
                    "modelo_b": r.pop("b_80plus"),
                },
                {
                    "label": "60–79%", "color": "#2563eb",
                    "total": r.pop("score_60_79"),
                    "modelo_a": r.pop("a_60_79"),
                    "modelo_b": r.pop("b_60_79"),
                },
                {
                    "label": "40–59%", "color": "#d97706",
                    "total": r.pop("score_40_59"),
                    "modelo_a": r.pop("a_40_59"),
                    "modelo_b": r.pop("b_40_59"),
                },
                {
                    "label": "< 40%",  "color": "#dc2626",
                    "total": r.pop("score_below_40"),
                    "modelo_a": r.pop("a_below_40"),
                    "modelo_b": r.pop("b_below_40"),
                },
            ]
            for bracket in r["funnel"]:
                bracket["pct"] = round(bracket["total"] * 100 / total, 1)

            return r
    finally:
        conn.close()


@router.get("/api/recommendations/product-lifecycle")
def get_product_lifecycle():
    """
    Top 50 produtos presentes no batch atual, por número de ofertas geradas.

    Para cada produto: rank, product_name, lifecycle_type (Consumível/Sazonal/Durável),
    avg_repurchase_days, repeat_rate_pct, total_buyers, n_offers, avg_score_pct.
    """
    conn = _get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""
                WITH latest_batch AS (
                    SELECT offer_batch_id
                    FROM   reco.offers
                    ORDER  BY generated_at DESC
                    LIMIT  1
                ),
                repurchase_intervals AS (
                    SELECT
                        product_id,
                        customer_id,
                        sale_date - LAG(sale_date) OVER (
                            PARTITION BY product_id, customer_id
                            ORDER BY sale_date
                        ) AS gap
                    FROM cur.order_items
                ),
                product_cycles AS (
                    SELECT
                        product_id,
                        ROUND(AVG(gap))::int AS avg_repurchase_days
                    FROM repurchase_intervals
                    WHERE gap IS NOT NULL
                    GROUP BY product_id
                ),
                product_buyers AS (
                    SELECT
                        product_id,
                        COUNT(DISTINCT customer_id)                                AS total_buyers,
                        COUNT(DISTINCT CASE WHEN cnt > 1 THEN customer_id END)     AS repeat_buyers
                    FROM (
                        SELECT product_id, customer_id, COUNT(*) AS cnt
                        FROM   cur.order_items
                        GROUP  BY product_id, customer_id
                    ) sub
                    GROUP BY product_id
                ),
                product_prices AS (
                    SELECT product_id, avg_unit_price
                    FROM cur.product_avg_price
                ),
                offer_stats AS (
                    SELECT
                        product_id,
                        COUNT(*)::int                AS n_offers,
                        ROUND(AVG(score) * 100)::int AS avg_score_pct
                    FROM reco.offers
                    WHERE offer_batch_id = (SELECT offer_batch_id FROM latest_batch)
                    GROUP BY product_id
                )
                SELECT
                    ROW_NUMBER() OVER (ORDER BY os.n_offers DESC)::int  AS rank,
                    p.product_id::int,
                    p.description                                        AS product_name,
                    COALESCE(pc.avg_repurchase_days, 0)::int             AS avg_repurchase_days,
                    CASE
                        WHEN COALESCE(pc.avg_repurchase_days, 0) = 0 THEN 'Sem histórico'
                        WHEN pc.avg_repurchase_days < 90              THEN 'Consumível'
                        WHEN pc.avg_repurchase_days < 365             THEN 'Sazonal'
                        ELSE                                               'Durável'
                    END                                                  AS lifecycle_type,
                    ROUND(
                        COALESCE(pb.repeat_buyers, 0) * 100.0
                        / NULLIF(pb.total_buyers, 0), 1
                    )::float                                             AS repeat_rate_pct,
                    COALESCE(pb.total_buyers, 0)::int                    AS total_buyers,
                    COALESCE(pp.avg_unit_price, 0)::float                AS avg_unit_price,
                    os.n_offers,
                    os.avg_score_pct
                FROM cur.products p
                JOIN offer_stats os         ON os.product_id = p.product_id
                LEFT JOIN product_cycles pc ON pc.product_id = p.product_id
                LEFT JOIN product_buyers pb ON pb.product_id = p.product_id
                LEFT JOIN product_prices pp ON pp.product_id = p.product_id
                WHERE p.active = TRUE
                ORDER BY os.n_offers DESC
                LIMIT 50
            """)
            rows = cur.fetchall()
            result = []
            for r in rows:
                row = dict(r)
                row["repeat_rate_pct"] = float(row["repeat_rate_pct"] or 0)
                row["avg_unit_price"]  = float(row["avg_unit_price"]  or 0)
                result.append(row)
            return result
    finally:
        conn.close()


@router.get("/api/recommendations/lifecycle")
def get_lifecycle(
    search: str = Query("", description="Filtro por nome do produto"),
    tier: str = Query("", description="Filtro por tier: short, medium, long"),
    limit: int = Query(500, le=5000),
    offset: int = Query(0, ge=0),
    user: dict = Depends(get_current_user_info),
):
    """
    Retorna dados de ciclo de vida de todos os produtos ativos.

    Consulta a materialized view reco.product_lifecycle juntamente com
    cur.products para obter nomes. Suporta busca por nome e filtro por tier.
    """
    conn = _get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT
                    pl.product_id,
                    cp.description AS product_name,
                    pl.avg_days_between_purchases,
                    pl.median_days_between_purchases,
                    pl.lifecycle_tier,
                    pl.sample_size,
                    pl.distinct_customers
                FROM reco.product_lifecycle pl
                JOIN cur.products cp
                    ON cp.product_id = pl.product_id
                   AND cp.source_system = 'sqlserver_gp'
                WHERE cp.active = TRUE
                  AND (%s = '' OR cp.description ILIKE '%%' || %s || '%%')
                  AND (%s = '' OR pl.lifecycle_tier = %s)
                ORDER BY pl.avg_days_between_purchases ASC
                LIMIT %s OFFSET %s
                """,
                (search, search, tier, tier, limit, offset),
            )
            rows = cur.fetchall()
            result = []
            for r in rows:
                row = dict(r)
                result.append({
                    "product_id": int(row["product_id"]),
                    "product_name": row["product_name"] or "",
                    "avg_days": float(row["avg_days_between_purchases"]) if row["avg_days_between_purchases"] is not None else 0,
                    "median_days": float(row["median_days_between_purchases"]) if row["median_days_between_purchases"] is not None else 0,
                    "tier": row["lifecycle_tier"] or "",
                    "sample_size": int(row["sample_size"]) if row["sample_size"] is not None else 0,
                    "distinct_customers": int(row["distinct_customers"]) if row["distinct_customers"] is not None else 0,
                })
            return result
    finally:
        conn.close()


@router.get("/api/recommendations/lifecycle/stats")
def get_lifecycle_stats(
    user: dict = Depends(get_current_user_info),
):
    """
    Retorna estatisticas agregadas do ciclo de vida dos produtos:
    total de produtos com lifecycle, distribuicao por tier.
    """
    conn = _get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT
                    COUNT(*)::int AS total,
                    COUNT(*) FILTER (WHERE pl.lifecycle_tier = 'short')::int  AS count_short,
                    COUNT(*) FILTER (WHERE pl.lifecycle_tier = 'medium')::int AS count_medium,
                    COUNT(*) FILTER (WHERE pl.lifecycle_tier = 'long')::int   AS count_long
                FROM reco.product_lifecycle pl
                JOIN cur.products cp
                    ON cp.product_id = pl.product_id
                   AND cp.source_system = 'sqlserver_gp'
                WHERE cp.active = TRUE
                """
            )
            row = cur.fetchone()
            if row is None:
                return {"total": 0, "count_short": 0, "count_medium": 0, "count_long": 0}
            return dict(row)
    finally:
        conn.close()


@router.get("/api/recommendations/feedback/summary")
def get_feedback_summary():
    """
    Metricas de conversao: taxa geral, por modelo, por faixa de score, por fonte.
    """
    conn = _get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""
                WITH outcomes AS (
                    SELECT
                        o.offer_id,
                        o.offer_batch_id,
                        o.strategy,
                        o.score,
                        oo.converted,
                        oo.conversion_source,
                        oo.total_value
                    FROM reco.offers o
                    LEFT JOIN reco.offer_outcomes oo ON oo.offer_id = o.offer_id
                )
                SELECT
                    COUNT(*)::int AS total_offers,
                    COUNT(converted)::int AS evaluated,
                    COUNT(*) FILTER (WHERE converted = TRUE)::int AS converted,
                    COUNT(*) FILTER (WHERE converted = FALSE)::int AS not_converted,
                    COUNT(*) - COUNT(converted) AS pending,
                    CASE WHEN COUNT(converted) > 0
                         THEN ROUND(COUNT(*) FILTER (WHERE converted = TRUE) * 100.0 / COUNT(converted), 2)
                         ELSE 0 END AS conversion_rate,
                    COALESCE(SUM(total_value) FILTER (WHERE converted = TRUE), 0)::float AS total_converted_value,
                    -- By model
                    COUNT(*) FILTER (WHERE strategy = 'modelo_a_ranker' AND converted = TRUE)::int AS modelo_a_converted,
                    COUNT(*) FILTER (WHERE strategy = 'modelo_a_ranker' AND converted IS NOT NULL)::int AS modelo_a_evaluated,
                    COUNT(*) FILTER (WHERE strategy = 'modelo_b_colaborativo' AND converted = TRUE)::int AS modelo_b_converted,
                    COUNT(*) FILTER (WHERE strategy = 'modelo_b_colaborativo' AND converted IS NOT NULL)::int AS modelo_b_evaluated,
                    -- By score range
                    COUNT(*) FILTER (WHERE score >= 0.8 AND converted = TRUE)::int AS high_score_converted,
                    COUNT(*) FILTER (WHERE score >= 0.8 AND converted IS NOT NULL)::int AS high_score_evaluated,
                    COUNT(*) FILTER (WHERE score >= 0.5 AND score < 0.8 AND converted = TRUE)::int AS mid_score_converted,
                    COUNT(*) FILTER (WHERE score >= 0.5 AND score < 0.8 AND converted IS NOT NULL)::int AS mid_score_evaluated,
                    COUNT(*) FILTER (WHERE score < 0.5 AND converted = TRUE)::int AS low_score_converted,
                    COUNT(*) FILTER (WHERE score < 0.5 AND converted IS NOT NULL)::int AS low_score_evaluated,
                    -- By source
                    COUNT(*) FILTER (WHERE conversion_source = 'automatic' AND converted = TRUE)::int AS auto_converted,
                    COUNT(*) FILTER (WHERE conversion_source = 'manual' AND converted = TRUE)::int AS manual_converted
                FROM outcomes
            """)
            row = cur.fetchone()
            if row is None:
                return {}
            result = dict(row)
            # Compute model-level rates
            result["modelo_a_rate"] = round(
                result["modelo_a_converted"] * 100.0 / result["modelo_a_evaluated"], 2
            ) if result["modelo_a_evaluated"] > 0 else 0
            result["modelo_b_rate"] = round(
                result["modelo_b_converted"] * 100.0 / result["modelo_b_evaluated"], 2
            ) if result["modelo_b_evaluated"] > 0 else 0
            result["high_score_rate"] = round(
                result["high_score_converted"] * 100.0 / result["high_score_evaluated"], 2
            ) if result["high_score_evaluated"] > 0 else 0
            result["mid_score_rate"] = round(
                result["mid_score_converted"] * 100.0 / result["mid_score_evaluated"], 2
            ) if result["mid_score_evaluated"] > 0 else 0
            result["low_score_rate"] = round(
                result["low_score_converted"] * 100.0 / result["low_score_evaluated"], 2
            ) if result["low_score_evaluated"] > 0 else 0
            return result
    finally:
        conn.close()


@router.get("/api/recommendations/feedback/runs")
def get_feedback_runs():
    """
    Historico de runs de feedback (cross-reference) em ordem cronologica reversa.
    """
    conn = _get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""
                SELECT run_id, offer_batch_id::text, started_at, finished_at,
                       status, conversion_window_days, offers_evaluated,
                       offers_converted, conversion_rate, triggered_by, error_message
                FROM reco.feedback_runs
                ORDER BY started_at DESC
                LIMIT 50
            """)
            rows = cur.fetchall()
            result = []
            for r in rows:
                row = dict(r)
                if row.get("started_at"):
                    row["started_at"] = row["started_at"].isoformat()
                if row.get("finished_at"):
                    row["finished_at"] = row["finished_at"].isoformat()
                if row.get("conversion_rate") is not None:
                    row["conversion_rate"] = float(row["conversion_rate"])
                result.append(row)
            return result
    finally:
        conn.close()


@router.post("/api/recommendations/feedback/run")
def trigger_feedback_run(
    batch_id: Optional[str] = Query(None),
    window_days: int = Query(30, ge=1, le=365),
    admin: dict = Depends(require_admin),
):
    """
    Dispara uma execucao manual de cross-reference (feedback loop).
    Requer role admin.
    """
    import sys
    from pathlib import Path
    _root = Path(__file__).resolve().parent.parent.parent.parent
    if str(_root) not in sys.path:
        sys.path.insert(0, str(_root))
    from ml.feedback_loop import run_cross_reference

    conn = _get_db()
    try:
        result = run_cross_reference(
            conn,
            batch_id=batch_id,
            window_days=window_days,
            dry_run=False,
            triggered_by=f"api:{admin['username']}",
        )
        return result
    finally:
        conn.close()


@router.get("/api/recommendations/offers/export-feedback")
def export_feedback_excel(
    batch_id: Optional[str] = Query(None),
    strategy: Optional[str] = Query(None),
):
    """
    Exporta planilha Excel (.xlsx) para feedback manual das ofertas.

    Colunas editaveis: Vendido (S/N), Data Venda, Observacoes.
    O telefone e exibido completo para uso operacional.
    """
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, PatternFill

    conn = _get_db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            if batch_id is None:
                cur.execute("""
                    SELECT offer_batch_id::text FROM reco.offers
                    ORDER BY generated_at DESC LIMIT 1
                """)
                row = cur.fetchone()
                if row is None:
                    raise HTTPException(status_code=404, detail="No offers found")
                batch_id = row["offer_batch_id"]

            params = [batch_id]
            strategy_filter = ""
            if strategy:
                strategy_filter = "AND o.strategy = %s"
                params.append(strategy)

            cur.execute(f"""
                SELECT
                    o.offer_id,
                    sc.name AS customer_name,
                    cp.description AS product_name,
                    COALESCE(pp.avg_unit_price, 0)::float AS avg_unit_price,
                    ROUND((o.score * 100)::numeric, 0)::int AS score_pct,
                    o.strategy,
                    lp.last_purchase_date,
                    NULLIF(
                        CASE
                            WHEN LENGTH(REGEXP_REPLACE(COALESCE(sc.mobile, sc.phone), '[^0-9]', '', 'g')) >= 6
                            THEN COALESCE(sc.mobile, sc.phone)
                            ELSE NULL
                        END,
                    '') AS contact,
                    CASE
                        WHEN wc.has_whatsapp = TRUE  THEN 'whatsapp'
                        WHEN wc.has_whatsapp = FALSE THEN 'mobile_no_wpp'
                        WHEN LENGTH(REGEXP_REPLACE(COALESCE(sc.mobile, sc.phone), '[^0-9]', '', 'g')) = 11
                         AND SUBSTRING(REGEXP_REPLACE(COALESCE(sc.mobile, sc.phone), '[^0-9]', '', 'g'), 3, 1) = '9'
                        THEN 'mobile'
                        WHEN LENGTH(REGEXP_REPLACE(COALESCE(sc.mobile, sc.phone), '[^0-9]', '', 'g')) >= 10
                        THEN 'landline'
                        ELSE 'unknown'
                    END AS phone_type
                FROM reco.offers o
                JOIN stg.customers sc ON sc.customer_id_src = o.customer_id AND sc.source_system = 'sqlserver_gp'
                JOIN cur.products cp ON cp.product_id = o.product_id AND cp.source_system = 'sqlserver_gp'
                LEFT JOIN reco.whatsapp_cache wc
                    ON wc.phone_number = '55' || REGEXP_REPLACE(COALESCE(sc.mobile, sc.phone), '[^0-9]', '', 'g')
                   AND wc.validated_at > now() - interval '30 days'
                LEFT JOIN cur.customer_product_last_purchase lp
                    ON lp.customer_id = o.customer_id AND lp.product_id = o.product_id
                LEFT JOIN cur.product_avg_price pp
                    ON pp.product_id = o.product_id
                WHERE o.offer_batch_id = %s::uuid {strategy_filter}
                ORDER BY o.rank ASC
            """, params)
            rows = cur.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Feedback Ofertas"

        headers = ["offer_id", "Cliente", "Produto", "Valor Unit.", "Score (%)",
                    "Modelo", "Ultima Compra", "Contato", "Tipo Telefone", "Vendido (S/N)", "Data Venda", "Observacoes"]
        ws.append(headers)

        # Style header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Hide offer_id column (column A)
        ws.column_dimensions['A'].hidden = True

        strategy_labels = {
            "modelo_a_ranker": "Modelo A",
            "modelo_b_colaborativo": "Modelo B",
        }

        for r in rows:
            row = dict(r)
            last_purchase = ""
            if row.get("last_purchase_date"):
                d = row["last_purchase_date"]
                last_purchase = d.strftime("%d/%m/%Y") if hasattr(d, "strftime") else str(d)

            phone_type_label = {
                "whatsapp": "WPP \u2713",
                "mobile": "Celular",
                "mobile_no_wpp": "Sem WPP",
                "landline": "Fixo",
            }.get(row.get("phone_type", ""), "")

            ws.append([
                int(row["offer_id"]),
                row.get("customer_name") or "",
                row.get("product_name") or "",
                float(row["avg_unit_price"]) if row.get("avg_unit_price") else 0,
                int(row["score_pct"]) if row.get("score_pct") is not None else 0,
                strategy_labels.get(row.get("strategy", ""), row.get("strategy", "")),
                last_purchase,
                _format_phone(row.get("contact")),
                phone_type_label,
                "",  # Vendido (S/N) — para preenchimento
                "",  # Data Venda — para preenchimento
                "",  # Observacoes — para preenchimento
            ])

        # Add data validation: dropdown S/N for "Vendido" column (J)
        dv = DataValidation(type="list", formula1='"S,N"', allow_blank=True)
        dv.error = "Apenas S ou N"
        dv.errorTitle = "Valor invalido"
        ws.add_data_validation(dv)
        dv.add(f"J2:J{len(rows) + 1}")

        # Set column widths
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 18
        ws.column_dimensions['I'].width = 16
        ws.column_dimensions['J'].width = 14
        ws.column_dimensions['K'].width = 14
        ws.column_dimensions['L'].width = 25

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"feedback_ofertas_{date.today().isoformat()}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    finally:
        conn.close()


@router.post("/api/recommendations/feedback/import")
async def import_feedback_excel(
    file: UploadFile = File(...),
    admin: dict = Depends(require_admin),
):
    """
    Importa feedback manual via planilha Excel (.xlsx).

    Espera o formato gerado por /offers/export-feedback:
    coluna A=offer_id, J=Vendido(S/N), K=Data Venda, L=Observacoes.
    Requer role admin.
    """
    from openpyxl import load_workbook

    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Only .xlsx files accepted")

    contents = await file.read()
    buf = io.BytesIO(contents)

    try:
        wb = load_workbook(buf, read_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid Excel file")

    ws = wb.active
    rows_list = list(ws.iter_rows(min_row=2, values_only=True))

    conn = _get_db()
    try:
        total = 0
        success = 0
        errors = []

        with conn.cursor() as cur:
            for i, row in enumerate(rows_list, start=2):
                if not row or len(row) < 10:
                    continue
                total += 1

                offer_id = row[0]
                vendido = str(row[9]).strip().upper() if row[9] else ""
                data_venda = row[10]
                observacoes = str(row[11]).strip() if row[11] and len(row) > 11 else ""

                if vendido not in ("S", "N"):
                    errors.append({"line": i, "error": f"Vendido deve ser S ou N, got '{vendido}'"})
                    continue

                if not offer_id:
                    errors.append({"line": i, "error": "offer_id vazio"})
                    continue

                # Verify offer exists
                cur.execute("SELECT 1 FROM reco.offers WHERE offer_id = %s", (int(offer_id),))
                if cur.fetchone() is None:
                    errors.append({"line": i, "error": f"offer_id {offer_id} nao encontrado"})
                    continue

                converted = vendido == "S"
                conversion_date = None
                if converted and data_venda:
                    if isinstance(data_venda, str):
                        from datetime import datetime as dt
                        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
                            try:
                                conversion_date = dt.strptime(data_venda.strip(), fmt).date()
                                break
                            except ValueError:
                                continue
                    elif hasattr(data_venda, 'date'):
                        conversion_date = data_venda.date() if hasattr(data_venda, 'date') else data_venda
                    else:
                        conversion_date = data_venda

                try:
                    cur.execute(
                        """
                        INSERT INTO reco.offer_outcomes
                            (offer_id, converted, conversion_date, conversion_source, notes, matched_by)
                        VALUES (%s, %s, %s, 'manual', %s, %s)
                        ON CONFLICT (offer_id, order_id_src) DO UPDATE
                            SET converted = EXCLUDED.converted,
                                conversion_date = EXCLUDED.conversion_date,
                                notes = EXCLUDED.notes,
                                matched_at = now()
                        """,
                        (int(offer_id), converted, conversion_date, observacoes or None,
                         f"excel_import:{admin['username']}"),
                    )
                    success += 1
                except Exception as e:
                    errors.append({"line": i, "error": str(e)[:200]})

        conn.commit()
        return {
            "total": total,
            "success": success,
            "errors_count": len(errors),
            "errors": errors[:50],  # Limit error output
        }
    finally:
        conn.close()
